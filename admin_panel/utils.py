"""
Admin Panel Utility Functions and Helpers

This module provides various utility functions to support admin panel operations
including pagination, exports, notifications, caching, and Indian-specific validations.
"""

import os
import io
import csv
import json
import hashlib
import logging
import tempfile
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
from PIL import Image, ImageOps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.mail import send_mail, EmailMultiAlternatives
from django.core.cache import cache
from django.template.loader import render_to_string
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Avg, Sum
from django.db import models
from django.contrib.auth.models import User

from accounts.models import GupShupUser
from posts.models import Post
from .models import AdminAction, AdminUser

# Configure logging
logger = logging.getLogger(__name__)

# Indian states and cities for validation
INDIAN_STATES = {
    'AN': 'Andaman and Nicobar Islands',
    'AP': 'Andhra Pradesh', 
    'AR': 'Arunachal Pradesh',
    'AS': 'Assam',
    'BR': 'Bihar',
    'CH': 'Chandigarh',
    'CT': 'Chhattisgarh',
    'DN': 'Dadra and Nagar Haveli',
    'DD': 'Daman and Diu',
    'DL': 'Delhi',
    'GA': 'Goa',
    'GJ': 'Gujarat',
    'HR': 'Haryana',
    'HP': 'Himachal Pradesh',
    'JK': 'Jammu and Kashmir',
    'JH': 'Jharkhand',
    'KA': 'Karnataka',
    'KL': 'Kerala',
    'LA': 'Ladakh',
    'LD': 'Lakshadweep',
    'MP': 'Madhya Pradesh',
    'MH': 'Maharashtra',
    'MN': 'Manipur',
    'ML': 'Meghalaya',
    'MZ': 'Mizoram',
    'NL': 'Nagaland',
    'OR': 'Odisha',
    'PY': 'Puducherry',
    'PB': 'Punjab',
    'RJ': 'Rajasthan',
    'SK': 'Sikkim',
    'TN': 'Tamil Nadu',
    'TG': 'Telangana',
    'TR': 'Tripura',
    'UP': 'Uttar Pradesh',
    'UT': 'Uttarakhand',
    'WB': 'West Bengal'
}

# Major Indian cities
MAJOR_INDIAN_CITIES = [
    'Mumbai', 'Delhi', 'Bangalore', 'Hyderabad', 'Ahmedabad', 'Chennai', 
    'Kolkata', 'Surat', 'Pune', 'Jaipur', 'Lucknow', 'Kanpur', 'Nagpur',
    'Patna', 'Indore', 'Thane', 'Bhopal', 'Visakhapatnam', 'Vadodara',
    'Firozabad', 'Ludhiana', 'Rajkot', 'Agra', 'Siliguri', 'Nashik',
    'Faridabad', 'Patiala', 'Ghaziabad', 'Kalyan', 'Dombivli', 'Howrah',
    'Ranchi', 'Allahabad', 'Coimbatore', 'Jabalpur', 'Gwalior'
]


class PaginationHelper:
    """Helper class for consistent pagination across admin views"""
    
    @staticmethod
    def paginate_queryset(queryset, request, per_page=25):
        """
        Paginate a queryset
        
        Args:
            queryset: Django queryset to paginate
            request: HTTP request object
            per_page: Items per page (default: 25)
            
        Returns:
            dict: Contains page object and pagination info
        """
        paginator = Paginator(queryset, per_page)
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        # Calculate page range for pagination links
        current_page = page_obj.number
        total_pages = paginator.num_pages
        
        # Show 5 pages around current page
        start_page = max(1, current_page - 2)
        end_page = min(total_pages, current_page + 2)
        
        page_range = range(start_page, end_page + 1)
        
        return {
            'page_obj': page_obj,
            'paginator': paginator,
            'page_range': page_range,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'total_count': paginator.count,
            'current_page': current_page,
            'total_pages': total_pages,
        }


class ExportHelper:
    """Helper class for exporting data in various formats"""
    
    @staticmethod
    def export_to_csv(data, filename, headers=None):
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries or list of lists
            filename: Output filename
            headers: Optional list of headers
            
        Returns:
            HttpResponse: CSV file response
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write BOM for Excel compatibility with UTF-8
        response.write('\ufeff'.encode('utf8'))
        
        if data:
            if headers:
                writer.writerow(headers)
            elif isinstance(data[0], dict):
                writer.writerow(data[0].keys())
            
            for row in data:
                if isinstance(row, dict):
                    writer.writerow(row.values())
                else:
                    writer.writerow(row)
        
        return response
    
    @staticmethod
    def export_to_excel(data, filename, sheet_name="Data", headers=None):
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries or list of lists
            filename: Output filename
            sheet_name: Excel sheet name
            headers: Optional list of headers
            
        Returns:
            HttpResponse: Excel file response
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if data:
            # Write headers
            if headers:
                header_row = headers
            elif isinstance(data[0], dict):
                header_row = list(data[0].keys())
            else:
                header_row = [f"Column {i+1}" for i in range(len(data[0]))]
            
            for col, header in enumerate(header_row, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            for row_idx, row in enumerate(data, 2):
                if isinstance(row, dict):
                    for col_idx, value in enumerate(row.values(), 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                else:
                    for col_idx, value in enumerate(row, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        
        return response
    
    @staticmethod
    def export_to_pdf(data, filename, title="Report", headers=None):
        """
        Export data to PDF format
        
        Args:
            data: List of dictionaries or list of lists
            filename: Output filename
            title: PDF title
            headers: Optional list of headers
            
        Returns:
            HttpResponse: PDF file response
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # Center alignment
        )
        
        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        if data:
            # Prepare table data
            if headers:
                table_data = [headers]
            elif isinstance(data[0], dict):
                table_data = [list(data[0].keys())]
            else:
                table_data = []
            
            # Add data rows
            for row in data:
                if isinstance(row, dict):
                    table_data.append(list(row.values()))
                else:
                    table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class NotificationHelper:
    """Helper class for sending admin notifications"""
    
    @staticmethod
    def send_admin_notification(subject, message, admin_users=None, email_template=None):
        """
        Send notification to admin users
        
        Args:
            subject: Email subject
            message: Plain text message
            admin_users: List of AdminUser objects (if None, sends to all super_admins)
            email_template: Optional HTML template name
            
        Returns:
            bool: Success status
        """
        try:
            if admin_users is None:
                admin_users = AdminUser.objects.filter(role='super_admin', is_active=True)
            
            recipient_emails = [admin.email for admin in admin_users if admin.email]
            
            if not recipient_emails:
                logger.warning("No admin email addresses found for notification")
                return False
            
            if email_template:
                html_message = render_to_string(email_template, {'message': message})
                msg = EmailMultiAlternatives(
                    subject=f"[GupShup Admin] {subject}",
                    body=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=recipient_emails
                )
                msg.attach_alternative(html_message, "text/html")
                msg.send()
            else:
                send_mail(
                    subject=f"[GupShup Admin] {subject}",
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=recipient_emails,
                    fail_silently=False,
                )
            
            logger.info(f"Admin notification sent to {len(recipient_emails)} admins")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send admin notification: {e}")
            return False
    
    @staticmethod
    def notify_critical_action(action, admin_user, details):
        """
        Send notification for critical admin actions
        
        Args:
            action: Action type (e.g., 'user_ban', 'bulk_delete')
            admin_user: AdminUser who performed the action
            details: Dictionary with action details
        """
        subject = f"Critical Action: {action.replace('_', ' ').title()}"
        
        message = f"""
Critical administrative action performed:

Action: {action}
Performed by: {admin_user.username} ({admin_user.get_role_display()})
Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} UTC
IP Address: {details.get('ip_address', 'Unknown')}

Details:
"""
        
        for key, value in details.items():
            if key != 'ip_address':
                message += f"  {key}: {value}\n"
        
        message += f"\nPlease review this action in the audit logs."
        
        NotificationHelper.send_admin_notification(
            subject, message, email_template='admin_panel/emails/critical_action.html'
        )


class CacheHelper:
    """Helper class for admin dashboard caching"""
    
    CACHE_TIMEOUT = 300  # 5 minutes
    
    @staticmethod
    def get_dashboard_stats(force_refresh=False):
        """
        Get cached dashboard statistics
        
        Args:
            force_refresh: If True, bypass cache and recalculate
            
        Returns:
            dict: Dashboard statistics
        """
        cache_key = 'admin_dashboard_stats'
        
        if not force_refresh:
            stats = cache.get(cache_key)
            if stats:
                return stats
        
        # Calculate fresh statistics
        stats = {
            'total_users': User.objects.count(),
            'active_users': User.objects.filter(is_active=True).count(),
            'total_posts': Post.objects.count(),
            'posts_today': Post.objects.filter(
                created_at__date=timezone.now().date()
            ).count(),
            'posts_this_week': Post.objects.filter(
                created_at__gte=timezone.now() - timedelta(days=7)
            ).count(),
            'posts_this_month': Post.objects.filter(
                created_at__gte=timezone.now() - timedelta(days=30)
            ).count(),
        }
        
        # Add user growth stats
        last_month = timezone.now() - timedelta(days=30)
        stats['new_users_this_month'] = User.objects.filter(
            date_joined__gte=last_month
        ).count()
        
        # Cache the results
        cache.set(cache_key, stats, CacheHelper.CACHE_TIMEOUT)
        
        return stats
    
    @staticmethod
    def clear_dashboard_cache():
        """Clear dashboard cache"""
        cache.delete('admin_dashboard_stats')
        cache.delete_pattern('admin_chart_*')
    
    @staticmethod
    def get_chart_data(chart_type, timeframe='7d'):
        """
        Get cached chart data
        
        Args:
            chart_type: Type of chart (users, posts, engagement)
            timeframe: Time frame (7d, 30d, 90d)
            
        Returns:
            dict: Chart data
        """
        cache_key = f'admin_chart_{chart_type}_{timeframe}'
        
        data = cache.get(cache_key)
        if data:
            return data
        
        # Generate fresh chart data based on type
        if chart_type == 'users':
            data = CacheHelper._generate_user_chart_data(timeframe)
        elif chart_type == 'posts':
            data = CacheHelper._generate_post_chart_data(timeframe)
        elif chart_type == 'engagement':
            data = CacheHelper._generate_engagement_chart_data(timeframe)
        else:
            data = {}
        
        # Cache for 15 minutes
        cache.set(cache_key, data, 900)
        
        return data
    
    @staticmethod
    def _generate_user_chart_data(timeframe):
        """Generate user growth chart data"""
        days = {'7d': 7, '30d': 30, '90d': 90}[timeframe]
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)
        
        data = {'labels': [], 'datasets': []}
        
        # Generate daily user counts
        current_date = start_date
        while current_date <= end_date:
            daily_count = User.objects.filter(
                date_joined__date=current_date
            ).count()
            
            data['labels'].append(current_date.strftime('%Y-%m-%d'))
            if not data['datasets']:
                data['datasets'].append({'label': 'New Users', 'data': []})
            data['datasets'][0]['data'].append(daily_count)
            
            current_date += timedelta(days=1)
        
        return data
    
    @staticmethod
    def _generate_post_chart_data(timeframe):
        """Generate post creation chart data"""
        days = {'7d': 7, '30d': 30, '90d': 90}[timeframe]
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)
        
        data = {'labels': [], 'datasets': []}
        
        current_date = start_date
        while current_date <= end_date:
            daily_count = Post.objects.filter(
                created_at__date=current_date
            ).count()
            
            data['labels'].append(current_date.strftime('%Y-%m-%d'))
            if not data['datasets']:
                data['datasets'].append({'label': 'Posts Created', 'data': []})
            data['datasets'][0]['data'].append(daily_count)
            
            current_date += timedelta(days=1)
        
        return data
    
    @staticmethod
    def _generate_engagement_chart_data(timeframe):
        """Generate engagement metrics chart data"""
        # This would need to be implemented based on your engagement tracking
        return {'labels': [], 'datasets': []}


class ImageHelper:
    """Helper class for image processing"""
    
    @staticmethod
    def process_avatar(image_file, size=(150, 150)):
        """
        Process user avatar for admin display
        
        Args:
            image_file: Uploaded image file
            size: Tuple of (width, height)
            
        Returns:
            Image: Processed PIL Image
        """
        try:
            image = Image.open(image_file)
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Create thumbnail while maintaining aspect ratio
            image = ImageOps.fit(image, size, Image.Resampling.LANCZOS)
            
            return image
            
        except Exception as e:
            logger.error(f"Error processing avatar: {e}")
            return None
    
    @staticmethod
    def create_thumbnail(image_path, size=(100, 100)):
        """
        Create thumbnail from image path
        
        Args:
            image_path: Path to image file
            size: Tuple of (width, height)
            
        Returns:
            str: Path to thumbnail or None if failed
        """
        try:
            if not os.path.exists(image_path):
                return None
            
            image = Image.open(image_path)
            
            # Create thumbnail
            thumbnail = ImageOps.fit(image, size, Image.Resampling.LANCZOS)
            
            # Generate thumbnail path
            base_name = os.path.splitext(os.path.basename(image_path))[0]
            thumb_path = os.path.join(
                settings.MEDIA_ROOT, 'thumbnails', f"{base_name}_thumb.jpg"
            )
            
            # Ensure thumbnail directory exists
            os.makedirs(os.path.dirname(thumb_path), exist_ok=True)
            
            # Save thumbnail
            thumbnail.save(thumb_path, 'JPEG', quality=85)
            
            return thumb_path
            
        except Exception as e:
            logger.error(f"Error creating thumbnail: {e}")
            return None


class IndianLocationValidator:
    """Helper class for Indian location validation"""
    
    @staticmethod
    def validate_phone_number(phone):
        """
        Validate Indian phone number
        
        Args:
            phone: Phone number string
            
        Returns:
            dict: {'valid': bool, 'formatted': str, 'error': str}
        """
        import re
        
        # Remove all non-digit characters
        digits = re.sub(r'\D', '', phone)
        
        # Check for valid Indian mobile patterns
        if len(digits) == 10 and digits[0] in '6789':
            return {'valid': True, 'formatted': f"+91{digits}", 'error': None}
        elif len(digits) == 12 and digits.startswith('91') and digits[2] in '6789':
            return {'valid': True, 'formatted': f"+{digits}", 'error': None}
        elif len(digits) == 13 and digits.startswith('091'):
            formatted = f"+91{digits[3:]}"
            return {'valid': True, 'formatted': formatted, 'error': None}
        else:
            return {
                'valid': False,
                'formatted': None,
                'error': 'Invalid Indian mobile number format'
            }
    
    @staticmethod
    def validate_pincode(pincode):
        """
        Validate Indian PIN code
        
        Args:
            pincode: PIN code string
            
        Returns:
            dict: {'valid': bool, 'error': str}
        """
        import re
        
        # Remove spaces and check format
        cleaned = re.sub(r'\s', '', str(pincode))
        
        if re.match(r'^[1-9][0-9]{5}$', cleaned):
            return {'valid': True, 'error': None}
        else:
            return {
                'valid': False,
                'error': 'Invalid PIN code format (should be 6 digits, not starting with 0)'
            }
    
    @staticmethod
    def get_state_choices():
        """Get choices for Indian states"""
        return [(code, name) for code, name in INDIAN_STATES.items()]
    
    @staticmethod
    def get_major_cities():
        """Get list of major Indian cities"""
        return MAJOR_INDIAN_CITIES
    
    @staticmethod
    def validate_state(state_code):
        """
        Validate Indian state code
        
        Args:
            state_code: Two-letter state code
            
        Returns:
            dict: {'valid': bool, 'name': str, 'error': str}
        """
        state_code = state_code.upper()
        
        if state_code in INDIAN_STATES:
            return {
                'valid': True,
                'name': INDIAN_STATES[state_code],
                'error': None
            }
        else:
            return {
                'valid': False,
                'name': None,
                'error': 'Invalid Indian state code'
            }


class HashtagAnalyzer:
    """Helper class for hashtag trend analysis"""
    
    @staticmethod
    def get_trending_hashtags(days=7, limit=20):
        """
        Get trending hashtags from posts
        
        Args:
            days: Number of days to analyze
            limit: Maximum number of hashtags to return
            
        Returns:
            list: List of hashtag dictionaries with counts
        """
        import re
        from collections import Counter
        
        # Get recent posts
        since_date = timezone.now() - timedelta(days=days)
        recent_posts = Post.objects.filter(
            created_at__gte=since_date
        ).values_list('content', flat=True)
        
        # Extract hashtags
        hashtag_pattern = r'#(\w+)'
        all_hashtags = []
        
        for content in recent_posts:
            if content:
                hashtags = re.findall(hashtag_pattern, content.lower())
                all_hashtags.extend(hashtags)
        
        # Count and sort
        hashtag_counts = Counter(all_hashtags)
        trending = [
            {'hashtag': f"#{tag}", 'count': count}
            for tag, count in hashtag_counts.most_common(limit)
        ]
        
        return trending
    
    @staticmethod
    def analyze_hashtag_sentiment(hashtag, days=7):
        """
        Analyze sentiment for a specific hashtag
        Uses basic sentiment analysis based on post engagement patterns
        
        Args:
            hashtag: Hashtag to analyze (with or without #)
            days: Number of days to analyze
            
        Returns:
            dict: Sentiment analysis results
        """
        hashtag = hashtag.lstrip('#').lower()
        since_date = timezone.now() - timedelta(days=days)
        
        # Get posts with this hashtag
        posts_with_hashtag = Post.objects.filter(
            content__icontains=f"#{hashtag}",
            created_at__gte=since_date
        )
        
        total_posts = posts_with_hashtag.count()
        
        # Basic sentiment analysis using engagement metrics
        if total_posts == 0:
            sentiment_scores = {'positive': 0.33, 'neutral': 0.33, 'negative': 0.34}
        else:
            # Calculate sentiment based on engagement patterns
            avg_likes = posts_with_hashtag.aggregate(
                likes=models.Avg('likes_count')
            )['likes'] or 0
            
            # Higher engagement generally indicates positive sentiment
            if avg_likes > 10:
                sentiment_scores = {'positive': 0.6, 'neutral': 0.3, 'negative': 0.1}
            elif avg_likes > 5:
                sentiment_scores = {'positive': 0.4, 'neutral': 0.4, 'negative': 0.2}
            else:
                sentiment_scores = {'positive': 0.3, 'neutral': 0.4, 'negative': 0.3}
        
        return {
            'hashtag': f"#{hashtag}",
            'total_posts': total_posts,
            'sentiment': sentiment_scores,
            'analysis_period': days
        }


def generate_secure_token(length=32):
    """
    Generate a secure random token
    
    Args:
        length: Token length
        
    Returns:
        str: Secure random token
    """
    import secrets
    return secrets.token_urlsafe(length)


def hash_sensitive_data(data):
    """
    Hash sensitive data for logging
    
    Args:
        data: String data to hash
        
    Returns:
        str: SHA-256 hash
    """
    return hashlib.sha256(str(data).encode()).hexdigest()


def format_file_size(size_bytes):
    """
    Format file size in human readable format
    
    Args:
        size_bytes: File size in bytes
        
    Returns:
        str: Formatted file size
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB", "TB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"


def get_client_ip(request):
    """
    Get client IP address from request
    
    Args:
        request: HTTP request object
        
    Returns:
        str: Client IP address
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def sanitize_filename(filename):
    """
    Sanitize filename for safe storage
    
    Args:
        filename: Original filename
        
    Returns:
        str: Sanitized filename
    """
    import re
    # Remove path separators and other potentially dangerous characters
    sanitized = re.sub(r'[^\w\-_\.]', '_', filename)
    # Limit length
    if len(sanitized) > 100:
        name, ext = os.path.splitext(sanitized)
        sanitized = name[:95] + ext
    return sanitized